# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import sys
from io import BytesIO

all_codecs = [
    'utf-8', 'gb2312', 'gbk', 'utf_16', 'ascii', 'big5', 'big5hkscs',
    'cp037', 'cp273', 'cp424', 'cp437',
    'cp500', 'cp720', 'cp737', 'cp775', 'cp850', 'cp852', 'cp855', 'cp856', 'cp857',
    'cp858', 'cp860', 'cp861', 'cp862', 'cp863', 'cp864', 'cp865', 'cp866', 'cp869',
    'cp874', 'cp875', 'cp932', 'cp949', 'cp950', 'cp1006', 'cp1026', 'cp1125',
    'cp1140', 'cp1250', 'cp1251', 'cp1252', 'cp1253', 'cp1254', 'cp1255', 'cp1256',
    'cp1257', 'cp1258', 'euc_jp', 'euc_jis_2004', 'euc_jisx0213', 'euc_kr',
    'gb2312', 'gb18030', 'hz', 'iso2022_jp', 'iso2022_jp_1', 'iso2022_jp_2',
    'iso2022_jp_2004', 'iso2022_jp_3', 'iso2022_jp_ext', 'iso2022_kr', 'latin_1',
    'iso8859_2', 'iso8859_3', 'iso8859_4', 'iso8859_5', 'iso8859_6', 'iso8859_7',
    'iso8859_8', 'iso8859_9', 'iso8859_10', 'iso8859_11', 'iso8859_13',
    'iso8859_14', 'iso8859_15', 'iso8859_16', 'johab', 'koi8_r', 'koi8_t', 'koi8_u',
    'kz1048', 'mac_cyrillic', 'mac_greek', 'mac_iceland', 'mac_latin2', 'mac_roman',
    'mac_turkish', 'ptcp154', 'shift_jis', 'shift_jis_2004', 'shift_jisx0213',
    'utf_32', 'utf_32_be', 'utf_32_le''utf_16_be', 'utf_16_le', 'utf_7'
]


def find_codec(blob):
    global all_codecs
    for c in all_codecs:
        try:
            blob[:1024].decode(c)
            return c
        except Exception as e:
            pass
        try:
            blob.decode(c)
            return c
        except Exception as e:
            pass

    return "utf-8"


class ExcelParser:
    def html(self, fnm, chunk_rows=256):
        if isinstance(fnm, str):
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(fnm))

        tb_chunks = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows: continue

            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{t.value}</th>"
            tb_rows_0 += "</tr>"

            for chunk_i in range((len(rows) - 1) // chunk_rows + 1):
                tb = ""
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                for r in list(rows[1 + chunk_i * chunk_rows:1 + (chunk_i + 1) * chunk_rows]):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{c.value}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks

    def __call__(self, fnm):
        if isinstance(fnm, str):
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(fnm))
        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows: continue
            ti = list(rows[0])
            for r in list(rows[1:]):
                l = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    l.append(t)
                l = "; ".join(l)
                if sheetname.lower().find("sheet") < 0:
                    l += " ——" + sheetname
                res.append(l)
        return res

    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = load_workbook(BytesIO(binary))
            total = 0
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                total += len(list(ws.rows))
                return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    psr = ExcelParser()
    print(psr.html("../data/excel/ex.xlsx"))
